import re
import openpyxl
from openpyxl.styles import Alignment
from datetime import datetime

file = open("poskakao.txt", "r", encoding='UTF8')
file_content = file.readlines()
pattern_general = r'\d{4}\s\w+\s일반' #일반 정규식
pattern_daily = r'\d{4}\s\w+\s데일리' #데일리 정규식
pattern_success = r'\d{4}\s\w+\s성공' #성공 정규식
pattern_fail = r'\d{4}\s\w+\s실패' #실패 정규식

matches_general = []
matches_daily = []
matches_success = []
matches_fail = []

nameSet = []
dateSet = []

for line in file_content:
    line_matches_daily = re.findall(pattern_daily, line) #데일리 수집
    line_matches_general = re.findall(pattern_general, line) #일반 수집
    line_matches_success = re.findall(pattern_success, line) #성공 수집
    line_matches_fail = re.findall(pattern_fail, line) #실패 수집

    matches_success.extend(line_matches_success)
    matches_fail.extend(line_matches_fail)
    matches_daily.extend(line_matches_daily)
    matches_general.extend(line_matches_general)

for match in matches_general:
    if match[5:8] not in nameSet: #match[5:8]은 이름 정보인데 nameSet에 없으면 추가함
        nameSet.append(match[5:8])

    formatted_date = f'{match[0:2]}월{match[2:4]}일'
    if formatted_date not in dateSet: #match[5:8]은 이름 정보인데 nameSet에 없으면 추가함
        dateSet.append(formatted_date)

print("성공:")
for match in matches_success:
    print(match) #없애도 됨

print("실패:")
for match in matches_fail:
    print(match) #없애도 됨

nameSet = sorted(nameSet) # 오름차순으로 정렬
print(nameSet) #없애도 됨
print(dateSet) #없애도 됨

file.close()

wb = openpyxl.Workbook()
ws = wb.active #첫번째 시트
wb.active.title = "test python"

nameRow = 66 #B부터 시작하는 이유는 A열에는 날짜를 넣어야 하기 때문에
dateCol = 2 #날짜 인덱스
dateColl = 3 #날짜 인덱스

for name in nameSet:
    ws[f'{chr(nameRow)}1'].value = name
    nameRow += 1

for date in dateSet:
    ws[f'A{dateCol}'].value = date

    dateMerge = f'A{dateCol}:A{dateColl}'
    ws.merge_cells(f'{dateMerge}')  # A1:A2 행을 합병

    dateCol += 2
    dateColl += 2

ws.column_dimensions['A'].width = len(str(ws['A2'].value))+2
from openpyxl.styles import Alignment

# 가운데 맞춤을 적용할 셀 범위
cell_range = f"A1:{chr(nameRow)}{dateColl}"
# 가운데 맞춤 스타일 생성
alignment = Alignment(horizontal="center", vertical="center")

# 셀 범위에 가운데 맞춤 스타일 적용
for row in ws[cell_range]:
    for cell in row:
        cell.alignment = alignment

wb.save(filename = "testpy.xlsx")